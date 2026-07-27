import io
import re
import unicodedata
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from ..models import Aluno, AluTurma, Aula, DocTurma, Materia, Professor, Turma

# O arquivo TS3 TURMA 2.xls usa treze colunas estreitas para as aulas e
# aproximadamente 29 alunos por página impressa.
DATAS_POR_PAGINA = 13
ALUNOS_POR_PAGINA = 29
COLUNA_PRIMEIRA_DATA = 4
COLUNA_ULTIMA_DATA = COLUNA_PRIMEIRA_DATA + DATAS_POR_PAGINA - 1
LINHAS_POR_PAGINA = 3 + ALUNOS_POR_PAGINA

FONTE_TABELA = "Palatino Linotype"
FONTE_NOMES = "Times New Roman"
BORDA_FINA = Side(style="thin", color="FF000000")
BORDA_MEDIA = Side(style="medium", color="FF000000")


def _nome_arquivo(valor: str) -> str:
    normalizado = unicodedata.normalize("NFKD", valor)
    sem_acentos = normalizado.encode("ascii", "ignore").decode("ascii")
    limpo = re.sub(r"[^A-Za-z0-9_-]+", "_", sem_acentos.strip())
    return limpo.strip("_") or "diario"


def _borda(
    esquerda: Side = BORDA_FINA,
    direita: Side = BORDA_FINA,
    topo: Side = BORDA_FINA,
    inferior: Side = BORDA_FINA,
) -> Border:
    return Border(left=esquerda, right=direita, top=topo, bottom=inferior)


def _aplicar_borda_mesclada(
    ws,
    linha: int,
    coluna_inicial: int,
    coluna_final: int,
) -> None:
    for coluna in range(coluna_inicial, coluna_final + 1):
        ws.cell(linha, coluna).border = Border(
            left=BORDA_MEDIA if coluna == coluna_inicial else Side(style=None),
            right=BORDA_MEDIA if coluna == coluna_final else Side(style=None),
            top=BORDA_MEDIA,
            bottom=BORDA_MEDIA,
        )


def _configurar_folha(
    ws,
    turma: Turma,
    materia: Materia,
    professor: Professor | None,
    ano: str,
) -> None:
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.sheet_properties.pageSetUpPr.autoPageBreaks = False
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = False
    ws.page_margins.left = 0.95
    ws.page_margins.right = 0.7
    ws.page_margins.top = 0.62
    ws.page_margins.bottom = 0.42
    ws.page_margins.header = 0.12
    ws.page_margins.footer = 0.12

    turma_nome = (turma.nome or str(turma.cod_tur)).strip()
    materia_nome = (materia.NOME or "").strip()
    professor_nome = (professor.nome or "").strip() if professor else "A definir"
    ws.oddHeader.center.text = (
        f"Diário de {materia_nome} - {turma_nome}\nProfº {professor_nome}"
    )
    ws.oddHeader.center.font = "Arial,Bold"
    ws.oddHeader.center.size = 11
    ws.oddFooter.center.text = f"Centro TOV / {ano} - {turma_nome}"
    ws.oddFooter.center.font = "Arial"
    ws.oddFooter.center.size = 9

    ws.column_dimensions["A"].width = 3.6
    ws.column_dimensions["B"].width = 6.6
    ws.column_dimensions["C"].width = 29
    for coluna in range(COLUNA_PRIMEIRA_DATA, COLUNA_ULTIMA_DATA + 1):
        ws.column_dimensions[ws.cell(1, coluna).column_letter].width = 7.55


def _cabecalho_bloco(
    ws,
    linha_inicial: int,
    turma: Turma,
    aulas: list[Aula],
) -> None:
    linha_turma = linha_inicial
    linha_datas = linha_inicial + 1
    linha_separadora = linha_inicial + 2

    ws.merge_cells(
        start_row=linha_turma,
        start_column=1,
        end_row=linha_turma,
        end_column=3,
    )
    ws.cell(linha_turma, 1, (turma.nome or str(turma.cod_tur)).upper())
    ws.cell(linha_turma, 1).font = Font(name=FONTE_TABELA, size=10, bold=True)
    ws.cell(linha_turma, 1).alignment = Alignment(
        horizontal="center", vertical="center"
    )
    _aplicar_borda_mesclada(ws, linha_turma, 1, 3)

    ws.merge_cells(
        start_row=linha_turma,
        start_column=COLUNA_PRIMEIRA_DATA,
        end_row=linha_turma,
        end_column=COLUNA_ULTIMA_DATA,
    )
    _aplicar_borda_mesclada(
        ws, linha_turma, COLUNA_PRIMEIRA_DATA, COLUNA_ULTIMA_DATA
    )

    ws.merge_cells(
        start_row=linha_datas,
        start_column=1,
        end_row=linha_datas,
        end_column=3,
    )
    _aplicar_borda_mesclada(ws, linha_datas, 1, 3)

    for deslocamento in range(DATAS_POR_PAGINA):
        coluna = COLUNA_PRIMEIRA_DATA + deslocamento
        cell = ws.cell(linha_datas, coluna)
        if deslocamento < len(aulas):
            cell.value = aulas[deslocamento].data
            cell.number_format = "d-mmm"
        cell.font = Font(name=FONTE_TABELA, size=10, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _borda(
            esquerda=BORDA_MEDIA if deslocamento == 0 else BORDA_FINA,
            direita=(
                BORDA_MEDIA
                if deslocamento == DATAS_POR_PAGINA - 1
                else BORDA_FINA
            ),
            topo=BORDA_FINA,
            inferior=BORDA_MEDIA,
        )

    ws.merge_cells(
        start_row=linha_separadora,
        start_column=1,
        end_row=linha_separadora,
        end_column=COLUNA_ULTIMA_DATA,
    )
    _aplicar_borda_mesclada(ws, linha_separadora, 1, COLUNA_ULTIMA_DATA)

    ws.row_dimensions[linha_turma].height = 15.75
    ws.row_dimensions[linha_datas].height = 15
    ws.row_dimensions[linha_separadora].height = 15


def _linhas_alunos(
    ws,
    linha_inicial: int,
    alunos_pagina: list[tuple[int, str]],
    numero_inicial: int,
) -> None:
    for posicao in range(ALUNOS_POR_PAGINA):
        linha = linha_inicial + posicao
        tem_aluno = posicao < len(alunos_pagina)
        numero = numero_inicial + posicao if tem_aluno else None
        matricula, nome = alunos_pagina[posicao] if tem_aluno else (None, None)
        valores = [numero, matricula, nome]

        for coluna, valor in enumerate(valores, start=1):
            cell = ws.cell(linha, coluna, valor)
            cell.font = Font(
                name=FONTE_NOMES if coluna == 3 else FONTE_TABELA,
                size=10,
            )
            cell.alignment = Alignment(
                horizontal="left" if coluna == 3 else "center",
                vertical="center",
            )
            cell.border = _borda(
                esquerda=BORDA_MEDIA if coluna == 1 else BORDA_FINA,
                direita=BORDA_FINA,
                inferior=(
                    BORDA_MEDIA
                    if posicao == ALUNOS_POR_PAGINA - 1
                    else BORDA_FINA
                ),
            )

        for coluna in range(COLUNA_PRIMEIRA_DATA, COLUNA_ULTIMA_DATA + 1):
            cell = ws.cell(linha, coluna)
            cell.font = Font(name=FONTE_TABELA, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _borda(
                direita=(
                    BORDA_MEDIA
                    if coluna == COLUNA_ULTIMA_DATA
                    else BORDA_FINA
                ),
                inferior=(
                    BORDA_MEDIA
                    if posicao == ALUNOS_POR_PAGINA - 1
                    else BORDA_FINA
                ),
            )
        ws.row_dimensions[linha].height = 15


def _montar_lista_presenca(
    workbook: Workbook,
    turma: Turma,
    materia: Materia,
    professor: Professor | None,
    alunos: list[tuple[int, str]],
    aulas: list[Aula],
    ano: str,
) -> None:
    grupos_datas = [
        aulas[inicio : inicio + DATAS_POR_PAGINA]
        for inicio in range(0, len(aulas), DATAS_POR_PAGINA)
    ]
    grupos_alunos = [
        alunos[inicio : inicio + ALUNOS_POR_PAGINA]
        for inicio in range(0, len(alunos), ALUNOS_POR_PAGINA)
    ] or [[]]

    pagina = 0
    for aulas_pagina in grupos_datas:
        for indice_alunos, alunos_pagina in enumerate(grupos_alunos):
            if pagina == 0:
                ws = workbook.active
                ws.title = "Lista de Presença"
            else:
                ws = workbook.create_sheet(f"Presença {pagina + 1}")
            _configurar_folha(ws, turma, materia, professor, ano)

            linha_inicial = 1
            _cabecalho_bloco(
                ws,
                linha_inicial,
                turma,
                aulas_pagina,
            )
            _linhas_alunos(
                ws,
                linha_inicial + 3,
                alunos_pagina,
                indice_alunos * ALUNOS_POR_PAGINA + 1,
            )
            pagina += 1
            ws.print_area = (
                f"A1:{get_column_letter(COLUNA_ULTIMA_DATA)}{LINHAS_POR_PAGINA}"
            )
            ws.sheet_view.zoomScale = 85
            ws.sheet_view.zoomScaleNormal = 85


def gerar_diario_xlsx(db: Session, docturma_id: int) -> tuple[bytes, str]:
    vinculo = db.get(DocTurma, docturma_id)
    if not vinculo:
        raise ValueError("Matéria da turma não encontrada")
    turma = db.get(Turma, vinculo.cod_tur)
    materia = db.get(Materia, vinculo.cod_mat)
    professor = db.get(Professor, vinculo.cod_pro) if vinculo.cod_pro else None
    if not turma or not materia:
        raise ValueError("Turma ou matéria não encontrada")

    aulas = list(
        db.scalars(
            select(Aula)
            .where(
                Aula.docturma_id == docturma_id,
                Aula.status != "CANCELADA",
            )
            .order_by(Aula.data, Aula.hora_inicio)
        )
    )
    if not aulas:
        raise ValueError("Cadastre ao menos uma aula desta matéria no calendário")

    alunos = list(
        db.execute(
            select(Aluno.cod_alu, Aluno.nome)
            .join(AluTurma, AluTurma.cod_alu == Aluno.cod_alu)
            .where(
                AluTurma.cod_tur == vinculo.cod_tur,
                (AluTurma.status.is_(None))
                | (~AluTurma.status.in_(["I", "INATIVO"])),
            )
            .order_by(Aluno.nome)
        )
    )

    ano = vinculo.Ano or str(aulas[0].data.year if aulas else date.today().year)
    workbook = Workbook()
    _montar_lista_presenca(
        workbook,
        turma,
        materia,
        professor,
        alunos,
        aulas,
        ano,
    )

    buffer = io.BytesIO()
    workbook.save(buffer)
    turma_arquivo = _nome_arquivo(turma.nome or str(turma.cod_tur))
    materia_arquivo = _nome_arquivo((materia.NOME or "").strip())
    nome = f"Diario_{turma_arquivo}_{materia_arquivo}.xlsx"
    return buffer.getvalue(), nome
