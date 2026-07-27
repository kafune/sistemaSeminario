import csv
import hashlib
import io
import json
import re
import unicodedata
from datetime import datetime

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from pydantic import BaseModel, Field, ValidationError
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from ..database import get_db
from ..models import ImportacaoGoogleForms, ItemImportacaoGoogleForms
from .integracoes import (
    PreCadastroGoogleForms,
    processar_pre_cadastro,
)

router = APIRouter(prefix="/importacoes", tags=["importações"])
LIMITE_ARQUIVO = 15 * 1024 * 1024


def _resumo(solicitacao: ImportacaoGoogleForms) -> dict:
    return {
        "id": solicitacao.id,
        "tipo": solicitacao.tipo or "IMPORTACAO",
        "status": solicitacao.status,
        "criados": solicitacao.criados or 0,
        "atualizados": solicitacao.atualizados or 0,
        "ja_cadastrados": solicitacao.ja_cadastrados or 0,
        "ja_processados": solicitacao.ja_processados or 0,
        "erros": solicitacao.erros or 0,
        "mensagem": solicitacao.mensagem,
    }


def _solicitar_google_forms(tipo: str, db: Session) -> dict:
    existente = db.scalar(
        select(ImportacaoGoogleForms)
        .where(
            ImportacaoGoogleForms.tipo == tipo,
            ImportacaoGoogleForms.status.in_(["PENDENTE", "PROCESSANDO"]),
        )
        .order_by(ImportacaoGoogleForms.id.desc())
        .limit(1)
    )
    if existente:
        return _resumo(existente)

    solicitacao = ImportacaoGoogleForms(
        tipo=tipo,
        status="PENDENTE",
        solicitada_em=datetime.now(),
        criados=0,
        atualizados=0,
        ja_cadastrados=0,
        ja_processados=0,
        erros=0,
    )
    db.add(solicitacao)
    db.commit()
    db.refresh(solicitacao)
    return _resumo(solicitacao)


@router.post("/google-forms")
def solicitar_importacao_google_forms(db: Session = Depends(get_db)):
    return _solicitar_google_forms("IMPORTACAO", db)


@router.post("/google-forms/previa")
def solicitar_previa_google_forms(db: Session = Depends(get_db)):
    return _solicitar_google_forms("PREVIA", db)


@router.get("/google-forms/{importacao_id}")
def obter_importacao_google_forms(
    importacao_id: int, db: Session = Depends(get_db)
):
    solicitacao = db.get(ImportacaoGoogleForms, importacao_id)
    if not solicitacao:
        raise HTTPException(404, "Solicitação de importação não encontrada")
    return _resumo(solicitacao)


@router.get("/google-forms/{importacao_id}/itens")
def listar_itens_google_forms(
    importacao_id: int,
    busca: str = "",
    db: Session = Depends(get_db),
):
    solicitacao = db.get(ImportacaoGoogleForms, importacao_id)
    if not solicitacao or solicitacao.tipo != "PREVIA":
        raise HTTPException(404, "Prévia da planilha não encontrada")
    if solicitacao.status != "PREVIA_PRONTA":
        raise HTTPException(409, "A prévia da planilha ainda não está pronta")

    consulta = select(ItemImportacaoGoogleForms).where(
        ItemImportacaoGoogleForms.importacao_id == importacao_id
    )
    if busca.strip():
        consulta = consulta.where(
            ItemImportacaoGoogleForms.nome.like(f"%{busca.strip()}%")
        )
    itens = list(
        db.scalars(
            consulta.order_by(ItemImportacaoGoogleForms.nome).limit(5000)
        )
    )
    return {
        "itens": [
            {
                "id": item.id,
                "nome": item.nome,
                "e_mail": item.e_mail,
                "telefone": item.telefone,
                "turma_interesse": item.turma_interesse,
            }
            for item in itens
        ]
    }


class SelecaoGoogleFormsInput(BaseModel):
    ids: list[int] = Field(min_length=1, max_length=5000)


@router.post("/google-forms/{importacao_id}/importar-selecao")
def importar_selecao_google_forms(
    importacao_id: int,
    dados: SelecaoGoogleFormsInput,
    db: Session = Depends(get_db),
):
    solicitacao = db.get(ImportacaoGoogleForms, importacao_id)
    if not solicitacao or solicitacao.tipo != "PREVIA":
        raise HTTPException(404, "Prévia da planilha não encontrada")
    if solicitacao.status != "PREVIA_PRONTA":
        raise HTTPException(409, "Esta prévia não está disponível para importação")

    ids = list(dict.fromkeys(dados.ids))
    itens = list(
        db.scalars(
            select(ItemImportacaoGoogleForms)
            .where(
                ItemImportacaoGoogleForms.importacao_id == importacao_id,
                ItemImportacaoGoogleForms.id.in_(ids),
            )
            .order_by(ItemImportacaoGoogleForms.nome)
        )
    )
    if len(itens) != len(ids):
        raise HTTPException(400, "A seleção contém pessoas inválidas")

    totais = {
        "criados": 0,
        "atualizados": 0,
        "ja_cadastrados": 0,
        "ja_processados": 0,
        "erros": 0,
    }
    mapa_acoes = {
        "pre_cadastro_criado": "criados",
        "pre_cadastro_atualizado": "atualizados",
        "ja_cadastrado": "ja_cadastrados",
        "ja_processado": "ja_processados",
    }
    mensagens_erro: list[str] = []
    for item in itens:
        try:
            payload = json.loads(item.payload_json)
            dados_aluno = PreCadastroGoogleForms.model_validate(payload)
            resultado = processar_pre_cadastro(
                dados_aluno,
                db,
                origem="IMPORTACAO_GOOGLE_SELETIVA",
            )
            totais[mapa_acoes[resultado["acao"]]] += 1
        except (json.JSONDecodeError, ValidationError, ValueError) as erro:
            db.rollback()
            totais["erros"] += 1
            mensagens_erro.append(f"{item.nome}: {erro}")

    for campo, valor in totais.items():
        setattr(solicitacao, campo, valor)
    solicitacao.status = "CONCLUIDA"
    solicitacao.concluida_em = datetime.now()
    solicitacao.mensagem = "; ".join(mensagens_erro[:5])[:255] or None
    db.execute(
        delete(ItemImportacaoGoogleForms).where(
            ItemImportacaoGoogleForms.importacao_id == importacao_id
        )
    )
    db.commit()
    return {
        "ok": True,
        **totais,
        "mensagem": solicitacao.mensagem,
    }


def _normalizar(valor) -> str:
    texto = unicodedata.normalize("NFKD", str(valor or ""))
    texto = texto.encode("ascii", "ignore").decode().lower()
    return re.sub(r"[^a-z0-9]+", " ", texto).strip()


def _campo_do_cabecalho(cabecalho) -> str | None:
    valor = _normalizar(cabecalho)
    aliases = {
        "nome": "nome",
        "nome completo": "nome",
        "qual a turma de interesse": "turma_interesse",
        "turma de interesse": "turma_interesse",
        "telefone": "telefone",
        "celular": "telefone",
        "e mail": "e_mail",
        "email": "e_mail",
        "rg": "rg",
        "cpf": "cpf",
        "escolaridade": "escolaridade",
        "igreja": "igreja",
        "igreja da qual e membro": "igreja",
        "endereco da igreja": "endereco_igreja",
        "local da igreja": "endereco_igreja",
        "endereco completo da igreja incluindo bairro e cidade": "endereco_igreja",
        "pastor": "nome_pastor",
        "nome do pastor": "nome_pastor",
        "curso anterior de teologia": "cur_teologicos",
        "nome do conjuge": "nome_conjuge",
        "conjuge": "nome_conjuge",
    }
    if valor in aliases:
        return aliases[valor]
    if valor.startswith("voce ja fez algum curso anterior de teologia"):
        return "cur_teologicos"
    if valor.startswith("seu conjuge participara junto"):
        return "nome_conjuge"
    return None


def _ler_matriz(nome_arquivo: str, conteudo: bytes) -> list[list]:
    nome = (nome_arquivo or "").lower()
    if nome.endswith(".csv"):
        try:
            texto = conteudo.decode("utf-8-sig")
        except UnicodeDecodeError:
            texto = conteudo.decode("latin-1")
        amostra = texto[:4096]
        separadores = {sep: amostra.count(sep) for sep in [";", ",", "\t"]}
        separador = max(separadores, key=separadores.get)
        return list(csv.reader(io.StringIO(texto), delimiter=separador))

    if nome.endswith(".xlsx"):
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
        planilha = workbook.worksheets[0]
        matriz = [list(linha) for linha in planilha.iter_rows(values_only=True)]
        workbook.close()
        return matriz

    if nome.endswith(".xls"):
        import xlrd

        workbook = xlrd.open_workbook(file_contents=conteudo)
        planilha = workbook.sheet_by_index(0)
        return [planilha.row_values(indice) for indice in range(planilha.nrows)]

    raise HTTPException(
        400, "Formato não suportado. Envie um arquivo .csv, .xlsx ou .xls"
    )


def _payloads_da_matriz(matriz: list[list]) -> tuple[list[dict], list[str]]:
    if not matriz:
        raise HTTPException(400, "A planilha está vazia")

    colunas: dict[int, str] = {}
    for indice, cabecalho in enumerate(matriz[0]):
        campo = _campo_do_cabecalho(cabecalho)
        if campo and campo not in colunas.values():
            colunas[indice] = campo
    if "nome" not in colunas.values():
        raise HTTPException(400, "Não foi encontrada uma coluna com o cabeçalho Nome")

    payloads: list[dict] = []
    erros: list[str] = []
    for numero_linha, linha in enumerate(matriz[1:], start=2):
        if not any(str(valor or "").strip() for valor in linha):
            continue
        dados = {
            campo: str(linha[indice] or "").strip()
            for indice, campo in colunas.items()
            if indice < len(linha)
        }
        if not dados.get("nome"):
            erros.append(f"Linha {numero_linha}: nome não informado")
            continue
        identidade = json.dumps(dados, ensure_ascii=False, sort_keys=True)
        dados["inscricao_id"] = hashlib.sha256(
            identidade.encode("utf-8")
        ).hexdigest()
        payloads.append(dados)
    return payloads, erros


@router.post("/arquivo")
def importar_arquivo(
    arquivo: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    conteudo = arquivo.file.read(LIMITE_ARQUIVO + 1)
    if len(conteudo) > LIMITE_ARQUIVO:
        raise HTTPException(400, "O arquivo excede o limite de 15 MB")

    matriz = _ler_matriz(arquivo.filename or "", conteudo)
    payloads, mensagens_erro = _payloads_da_matriz(matriz)
    if not payloads and not mensagens_erro:
        raise HTTPException(400, "A planilha não possui alunos para importar")

    totais = {
        "criados": 0,
        "atualizados": 0,
        "ja_cadastrados": 0,
        "ja_processados": 0,
        "erros": len(mensagens_erro),
    }
    mapa_acoes = {
        "pre_cadastro_criado": "criados",
        "pre_cadastro_atualizado": "atualizados",
        "ja_cadastrado": "ja_cadastrados",
        "ja_processado": "ja_processados",
    }

    for numero_linha, payload in enumerate(payloads, start=2):
        try:
            dados = PreCadastroGoogleForms.model_validate(payload)
            resultado = processar_pre_cadastro(
                dados, db, origem="IMPORTACAO_ARQUIVO"
            )
            totais[mapa_acoes[resultado["acao"]]] += 1
        except (ValidationError, ValueError) as erro:
            db.rollback()
            totais["erros"] += 1
            mensagens_erro.append(f"Linha {numero_linha}: {erro}")

    return {
        "ok": True,
        **totais,
        "mensagem": "; ".join(mensagens_erro[:5]) or None,
    }
