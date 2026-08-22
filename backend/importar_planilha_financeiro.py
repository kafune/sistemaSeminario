"""Importa a planilha de controle de matrículas e mensalidades.

A planilha que a secretaria mantém à mão tem duas listas — alunos regulares e
transferidos — e marca com ``C`` o casal, cujas células de valor ficam mescladas
entre as duas linhas porque **o casal paga junto**. As pessoas estão misturadas
entre as turmas, então o script **casa cada nome com o aluno já cadastrado** e
respeita a turma em que ele está; ninguém é movido de turma pela importação.

O que ele faz, uma vez casados os nomes:

* aplica o plano (matrícula + mensalidades) a **cada turma** que tenha aluno na
  planilha;
* o cônjuge (a segunda linha do par) recebe o desconto percentual, que abate
  matrícula e mensalidade;
* o transferido recebe a condição de transferência, dispensado da matrícula;
* o que já foi pago entra como baixa, quitando as cobranças em ordem de
  vencimento — no casal, primeiro as do titular e depois as do cônjuge.

Ao final confere o resultado contra os valores da própria planilha. Se não
bater, desfaz tudo.

Uso, a partir da pasta backend/ com o .env configurado:

    # relatório de correspondência de nomes; não grava nada
    python importar_planilha_financeiro.py --arquivo ../PLANILHA.xlsx

    # grava o que casou exatamente
    python importar_planilha_financeiro.py --arquivo ../PLANILHA.xlsx --aplicar

    # aceita também os nomes parecidos e cria quem não existe
    python importar_planilha_financeiro.py --arquivo ../PLANILHA.xlsx \\
        --aceitar-aproximados --criar-novos --turma-novos 1 --aplicar
"""

import argparse
import re
import sys
import unicodedata
from datetime import date
from decimal import Decimal
from difflib import SequenceMatcher

from sqlalchemy import func, select

from app.database import Base, SessionLocal, engine
from app.models import (
    Aluno,
    AluTurma,
    Cobranca,
    CondicaoFinanceiraAluno,
    Pagamento,
    PlanoFinanceiro,
    TransacaoBancaria,
    Turma,
)
from app.schema import atualizar_schema
from app.services import financeiro as servico
from app.services.matriculas import sincronizar_matricula

ORIGEM = "PLANILHA"
# Abaixo disso não é nome parecido, é outro aluno.
SEMELHANCA_MINIMA = 0.88
# Partículas não distinguem ninguém: entram e saem conforme quem digitou.
PARTICULAS = frozenset({"DA", "DAS", "DE", "DEL", "DI", "DO", "DOS", "E", "VAN", "VON"})


# ---- nomes -----------------------------------------------------------------

def normalizar(nome: str) -> str:
    sem_acento = unicodedata.normalize("NFKD", nome or "")
    sem_acento = "".join(c for c in sem_acento if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", sem_acento).strip().upper()


def chave_curta(nome: str) -> str:
    """Primeiro nome + último sobrenome, sem as partículas do meio.

    "Marcos Antonio de Lima Filho" e "Marcos Antônio Lima Filho" caem na mesma
    chave; é o erro de digitação mais comum entre a planilha e o cadastro.
    """
    partes = tokens_significativos(nome)
    if not partes:
        return normalizar(nome)
    if len(partes) == 1:
        return partes[0]
    return f"{partes[0]} {partes[-1]}"


def tokens_significativos(nome: str) -> list[str]:
    """Nome sem as partículas: "de", "da", "dos" não distinguem ninguém."""
    return [parte for parte in normalizar(nome).split(" ") if parte and parte not in PARTICULAS]


def assinatura(nome: str) -> str:
    """Tokens significativos em ordem alfabética.

    Faz "Maria de Souza" e "Maria Souza" caírem na mesma chave — a partícula
    é ruído do mesmo tipo que o acento, não outro nome.
    """
    return " ".join(sorted(tokens_significativos(nome)))


def nome_contido(um: str, outro: str) -> bool:
    """Verdadeiro quando um nome é o outro pela metade.

    É o caso que a planilha cria o tempo todo: "Evaneide Maria" escrito à mão
    para quem está cadastrado como "Evaneide Maria da Silva Santos". Exige o
    mesmo primeiro nome e pelo menos dois tokens, senão "Maria" casaria com
    meia escola.
    """
    aqui, la = tokens_significativos(um), tokens_significativos(outro)
    if len(aqui) < 2 or len(la) < 2:
        return False
    curto, longo = (aqui, la) if len(aqui) <= len(la) else (la, aqui)
    if curto[0] != longo[0]:
        return False
    return all(parte in longo for parte in curto)


def semelhanca(um: str, outro: str) -> float:
    return SequenceMatcher(None, normalizar(um), normalizar(outro)).ratio()


# ---- leitura da planilha ---------------------------------------------------

def _decimal(valor) -> Decimal:
    """Converte a célula em dinheiro; ``PG`` e vazio viram zero."""
    if valor is None:
        return servico.ZERO
    texto = str(valor).strip().replace("R$", "").replace(".", "").replace(",", ".")
    if not texto or not re.fullmatch(r"-?\d+(\.\d+)?", texto):
        return servico.ZERO
    return servico.dinheiro(Decimal(texto))


def ler_planilha(caminho) -> tuple[list[dict], dict]:
    """Devolve as pessoas na ordem da planilha e os totais que ela declara."""
    try:
        import openpyxl
    except ImportError:  # pragma: no cover - dependência declarada
        sys.exit("openpyxl não instalado: pip install -r requirements.txt")

    livro = openpyxl.load_workbook(caminho, data_only=True)
    aba = livro[livro.sheetnames[0]]

    # A célula da coluna C mesclada em duas linhas é a marca do casal: uma
    # linha de valores para duas pessoas.
    titulares_de_casal = {
        intervalo.min_row
        for intervalo in aba.merged_cells.ranges
        if intervalo.min_col == 3 and intervalo.max_row == intervalo.min_row + 1
    }

    pessoas: list[dict] = []
    totais = {"pago": None, "a_pagar": None}
    bloco = "REGULAR"
    for indice, linha in enumerate(aba.iter_rows(values_only=True)):
        excel = indice + 1
        primeira = str(linha[0] or "").strip()
        nome = str(linha[1] or "").strip()
        if "TRANSFER" in normalizar(primeira) or "TRANSFER" in normalizar(nome):
            bloco = "TRANSFERENCIA"
            continue
        rotulo = normalizar(nome)
        if rotulo.startswith("VALOR TOTAL PAGO"):
            totais["pago"] = _decimal(linha[3])
            continue
        if rotulo.startswith("VALOR TOTAL A PAGAR"):
            totais["a_pagar"] = _decimal(linha[6])
            continue
        if not nome or not primeira.isdigit():
            continue
        pessoas.append(
            {
                "linha": excel,
                "nome": re.sub(r"\s+", " ", nome),
                "bloco": bloco,
                "titular_casal": excel in titulares_de_casal,
                "conjuge": (excel - 1) in titulares_de_casal,
                "pago": _decimal(linha[3]),
                "a_pagar": _decimal(linha[6]),
                "conta": str(linha[7] or "").strip() or None,
                "recibo": str(linha[8] or "").strip() or None,
            }
        )
    return pessoas, totais


# ---- correspondência com o cadastro ---------------------------------------

def casar_nomes(db, pessoas: list[dict]) -> list[dict]:
    """Liga cada linha da planilha ao aluno já cadastrado com aquele nome.

    São quatro camadas, da mais para a menos certa: nome idêntico; mesmos
    tokens em outra ordem ou sem as partículas; nome incompleto contido no
    completo; e finalmente parecido por semelhança. A primeira que encontrar um
    único candidato decide.

    Devolve, para cada pessoa, a situação da correspondência:

    ``EXATO``       mesmo nome, ignorando acento, caixa, partícula e ordem;
    ``PARCIAL``     a planilha tem o nome pela metade — precisa de confirmação;
    ``APROXIMADO``  um único parecido — precisa de confirmação;
    ``AMBIGUO``     mais de um candidato, então o script não escolhe;
    ``NOVO``        ninguém parecido no cadastro.
    """
    cadastro = list(db.scalars(select(Aluno).where(Aluno.nome.is_not(None))))
    por_nome: dict[str, list[Aluno]] = {}
    por_assinatura: dict[str, list[Aluno]] = {}
    por_chave: dict[str, list[Aluno]] = {}
    for aluno in cadastro:
        por_nome.setdefault(normalizar(aluno.nome), []).append(aluno)
        por_assinatura.setdefault(assinatura(aluno.nome), []).append(aluno)
        por_chave.setdefault(chave_curta(aluno.nome), []).append(aluno)

    correspondencias = []
    for pessoa in pessoas:
        nome = pessoa["nome"]
        camadas = [
            ("EXATO", por_nome.get(normalizar(nome), [])),
            ("EXATO", por_assinatura.get(assinatura(nome), [])),
            ("PARCIAL", [a for a in cadastro if nome_contido(nome, a.nome)]),
            ("APROXIMADO", por_chave.get(chave_curta(nome), [])),
            (
                "APROXIMADO",
                [a for a in cadastro if semelhanca(nome, a.nome) >= SEMELHANCA_MINIMA],
            ),
        ]
        for situacao, candidatos in camadas:
            if not candidatos:
                continue
            if len(candidatos) > 1:
                correspondencias.append(
                    {**pessoa, "situacao": "AMBIGUO", "candidatos": candidatos}
                )
                break
            correspondencias.append(
                {
                    **pessoa,
                    "situacao": situacao,
                    "aluno": candidatos[0],
                    "semelhanca": semelhanca(nome, candidatos[0].nome),
                }
            )
            break
        else:
            correspondencias.append({**pessoa, "situacao": "NOVO"})
    return correspondencias


def turma_do_aluno(db, aluno: Aluno) -> int | None:
    if aluno.cod_tur is not None:
        return aluno.cod_tur
    return db.scalar(select(AluTurma.cod_tur).where(AluTurma.cod_alu == aluno.cod_alu))


def selecionar(
    db,
    correspondencias: list[dict],
    *,
    aceitar_aproximados: bool,
    criar_novos: bool,
    turma_novos: int | None,
) -> tuple[list[dict], list[tuple[dict, str]]]:
    """Separa quem entra na importação de quem fica de fora, e por quê."""
    entram: list[dict] = []
    ficam_de_fora: list[tuple[dict, str]] = []
    for item in correspondencias:
        situacao = item["situacao"]
        if situacao == "AMBIGUO":
            nomes = ", ".join(f"{a.nome} (#{a.cod_alu})" for a in item["candidatos"][:3])
            ficam_de_fora.append((item, f"mais de um aluno com esse nome ({nomes})"))
            continue
        if situacao in ("APROXIMADO", "PARCIAL") and not aceitar_aproximados:
            como = "parecido com" if situacao == "APROXIMADO" else "nome incompleto de"
            ficam_de_fora.append(
                (item, f"{como} {item['aluno'].nome!r}; use --aceitar-aproximados")
            )
            continue
        if situacao == "NOVO":
            if not criar_novos:
                ficam_de_fora.append((item, "não está no cadastro; use --criar-novos"))
                continue
            if turma_novos is None:
                ficam_de_fora.append((item, "aluno novo sem --turma-novos"))
                continue
            entram.append({**item, "cod_tur": turma_novos})
            continue

        cod_tur = turma_do_aluno(db, item["aluno"])
        if cod_tur is None:
            if turma_novos is None:
                ficam_de_fora.append((item, "aluno sem turma; informe --turma-novos"))
                continue
            cod_tur = turma_novos
        entram.append({**item, "cod_tur": cod_tur})
    return entram, ficam_de_fora


# ---- gravação --------------------------------------------------------------

def garantir_plano(
    db,
    cod_tur: int,
    *,
    matricula: Decimal,
    mensalidade: Decimal,
    parcelas: int,
    primeira_mensalidade: date,
    dia_vencimento: int,
) -> PlanoFinanceiro:
    plano = db.scalar(select(PlanoFinanceiro).where(PlanoFinanceiro.cod_tur == cod_tur))
    if plano is None:
        plano = PlanoFinanceiro(cod_tur=cod_tur)
        db.add(plano)
    plano.valor_matricula = servico.dinheiro(matricula)
    plano.valor_mensalidade = servico.dinheiro(mensalidade)
    plano.parcelas = parcelas
    plano.dia_vencimento = dia_vencimento
    plano.primeira_mensalidade = primeira_mensalidade
    plano.vencimento_matricula = primeira_mensalidade
    plano.observacao = "Importado da planilha de controle de matrículas."
    plano.atualizado_por = ORIGEM
    db.flush()
    return plano


def definir_condicao(
    db,
    item: dict,
    aluno: Aluno,
    cod_tur: int,
    *,
    desconto_conjuge: Decimal,
    parcelas_transferencia: int | None,
) -> CondicaoFinanceiraAluno | None:
    """Traduz as marcas da planilha (C e o bloco) na condição do aluno."""
    transferencia = item["bloco"] == "TRANSFERENCIA"
    conjuge = item["conjuge"]
    if not transferencia and not conjuge:
        return None

    condicao = db.scalar(
        select(CondicaoFinanceiraAluno).where(
            CondicaoFinanceiraAluno.cod_tur == cod_tur,
            CondicaoFinanceiraAluno.cod_alu == aluno.cod_alu,
        )
    )
    if condicao is None:
        condicao = CondicaoFinanceiraAluno(cod_alu=aluno.cod_alu, cod_tur=cod_tur)
        db.add(condicao)
    # O tipo é sempre explícito: o cônjuge do bloco regular não é transferido.
    condicao.tipo = "TRANSFERENCIA" if transferencia else "REGULAR"
    if transferencia:
        condicao.cobra_matricula = "N"
        if parcelas_transferencia is not None:
            condicao.parcelas = parcelas_transferencia
        condicao.observacao = (
            "Transferido de outro seminário; dispensado da matrícula (planilha). "
            "Confirme quantos módulos ele ainda vai cursar."
        )
    if conjuge:
        condicao.desconto_percentual = servico.dinheiro(desconto_conjuge)
        condicao.desconto_motivo = "Casal — cônjuge com desconto"
        condicao.desconto_na_matricula = "S"
    condicao.atualizado_por = ORIGEM
    db.flush()
    return condicao


def quitar(
    db,
    item: dict,
    alunos_do_pagamento: list[Aluno],
    *,
    data_pagamento: date,
    relatar=print,
) -> Decimal:
    """Aplica o valor pago às cobranças, na ordem em que vencem.

    No casal o dinheiro entrou junto: começa pelo titular e transborda para o
    cônjuge, que é como a planilha registra — um depósito, duas pessoas.
    """
    # O que a planilha diz que entrou, menos o que uma rodada anterior já
    # lançou: sem isso, importar duas vezes pagaria a mesma conta duas vezes.
    ja_importado = servico.dinheiro(
        db.scalar(
            select(func.sum(Pagamento.valor))
            .join(Cobranca, Cobranca.id == Pagamento.cobranca_id)
            .where(
                Pagamento.registrado_por == ORIGEM,
                Cobranca.cod_alu.in_([a.cod_alu for a in alunos_do_pagamento]),
            )
        )
    )
    restante = item["pago"] - ja_importado
    if restante <= servico.ZERO:
        return servico.ZERO
    observacao = " · ".join(
        parte for parte in ("Importado da planilha", item["conta"], item["recibo"]) if parte
    )
    aplicado = servico.ZERO
    for aluno in alunos_do_pagamento:
        cobrancas = db.scalars(
            select(Cobranca)
            .where(Cobranca.cod_alu == aluno.cod_alu, Cobranca.status == "ABERTA")
            .order_by(Cobranca.vencimento, Cobranca.tipo.desc(), Cobranca.parcela)
        )
        for cobranca in cobrancas:
            if restante <= servico.ZERO:
                break
            saldo = servico.dinheiro(cobranca.valor) - servico.total_pago(db, cobranca.id)
            if saldo <= servico.ZERO:
                continue
            valor = min(saldo, restante)
            servico.registrar_pagamento(
                db,
                cobranca,
                valor=valor,
                data_pagamento=data_pagamento,
                forma="TRANSFERENCIA",
                observacao=observacao,
                registrado_por=ORIGEM,
            )
            restante -= valor
            aplicado += valor
    if restante > servico.ZERO:
        relatar(f"  ! {item['nome']}: sobraram R$ {restante} sem cobrança para quitar")
    return aplicado


# ---- conferência -----------------------------------------------------------

def _foto(db, aluno: Aluno) -> tuple[Decimal, Decimal]:
    """Matrícula + primeira mensalidade: o recorte que a planilha retrata."""
    cobrancas = list(
        db.scalars(
            select(Cobranca).where(
                Cobranca.cod_alu == aluno.cod_alu,
                Cobranca.tipo.in_(("MATRICULA", "MENSALIDADE")),
                Cobranca.parcela == 1,
            )
        )
    )
    pagos = servico.pagos_por_cobranca(db, [c.id for c in cobrancas])
    pago = sum((pagos.get(c.id, servico.ZERO) for c in cobrancas), servico.ZERO)
    aberto = sum(
        (
            max(servico.dinheiro(c.valor) - pagos.get(c.id, servico.ZERO), servico.ZERO)
            for c in cobrancas
        ),
        servico.ZERO,
    )
    return pago, aberto


def conferir(db, importados: list[dict], totais: dict, *, relatar=print) -> bool:
    """Compara o que ficou no banco com o que a planilha diz de cada aluno.

    A comparação é sobre quem entrou: se alguém ficou de fora, o total geral da
    planilha naturalmente não fecha, e é o relatório de exclusões que explica.
    """
    por_linha = {item["linha"]: item for item in importados}
    pago_banco = aberto_banco = servico.ZERO
    pago_planilha = aberto_planilha = servico.ZERO
    divergencias = []

    for item in importados:
        pago, aberto = _foto(db, item["aluno"])
        pago_banco += pago
        aberto_banco += aberto
        if item["conjuge"]:
            continue  # a planilha lança o casal na linha do titular

        pago_planilha += item["pago"]
        aberto_planilha += item["a_pagar"]
        esperado_pago, esperado_aberto = item["pago"], item["a_pagar"]
        if item["titular_casal"]:
            conjuge = por_linha.get(item["linha"] + 1)
            if conjuge is None:
                # Sem o cônjuge no lote a linha não tem como fechar.
                divergencias.append(
                    f"  ! {item['nome']}: cônjuge fora da importação; confira à mão"
                )
                continue
            pago_conjuge, aberto_conjuge = _foto(db, conjuge["aluno"])
            pago += pago_conjuge
            aberto += aberto_conjuge
        if pago != esperado_pago or aberto != esperado_aberto:
            divergencias.append(
                f"  ! {item['nome']}: pago {pago} (planilha {esperado_pago}), "
                f"em aberto {aberto} (planilha {esperado_aberto})"
            )

    relatar("")
    relatar("Conferência contra a planilha (sobre quem entrou)")
    relatar(f"  pago     no banco {pago_banco:>9}   planilha {pago_planilha:>9}")
    relatar(f"  a pagar  no banco {aberto_banco:>9}   planilha {aberto_planilha:>9}")
    if totais.get("pago") is not None:
        relatar(f"  (a planilha inteira declara pago {totais['pago']} e a pagar {totais['a_pagar']})")
    for linha in divergencias:
        relatar(linha)
    bate = not divergencias and pago_banco == pago_planilha and aberto_banco == aberto_planilha
    relatar("  => " + ("confere" if bate else "NÃO confere"))
    return bate


# ---- diagnóstico e limpeza -------------------------------------------------

def cobrancas_do_aluno(db, cod_alu: int) -> list[Cobranca]:
    return list(
        db.scalars(
            select(Cobranca)
            .where(Cobranca.cod_alu == cod_alu)
            .order_by(Cobranca.cod_tur, Cobranca.vencimento, Cobranca.id)
        )
    )


def diagnosticar(db, correspondencias: list[dict], relatar=print) -> dict:
    """Mostra o que já existe no banco para cada aluno da planilha.

    Serve para responder por que a conferência não fechou: quase sempre é
    cobrança ou baixa que veio de antes — de uma importação anterior, de um
    lançamento na tela ou de uma geração feita pelo plano da turma.
    """
    nomes_turma = {cod: nome for cod, nome in db.execute(select(Turma.cod_tur, Turma.nome))}
    resumo = {
        "com_cobranca": 0,
        "em_mais_de_uma_turma": 0,
        "com_pagamento": 0,
        "por_origem": {},
        "alunos_em_duas_turmas": [],
    }
    relatar("")
    relatar("O que já existe no banco para estes alunos")
    relatar("-" * 96)
    for item in correspondencias:
        aluno = item.get("aluno")
        if aluno is None:
            continue
        cobrancas = cobrancas_do_aluno(db, aluno.cod_alu)
        if not cobrancas:
            continue
        resumo["com_cobranca"] += 1
        pagos = servico.pagos_por_cobranca(db, [c.id for c in cobrancas])
        turmas = sorted({c.cod_tur for c in cobrancas if c.cod_tur is not None})
        if len(turmas) > 1:
            resumo["em_mais_de_uma_turma"] += 1
            resumo["alunos_em_duas_turmas"].append(aluno.nome)

        pagamentos = list(
            db.scalars(
                select(Pagamento).where(Pagamento.cobranca_id.in_([c.id for c in cobrancas]))
            )
        )
        total_pago = sum((servico.dinheiro(p.valor) for p in pagamentos), servico.ZERO)
        if pagamentos:
            resumo["com_pagamento"] += 1
        for pagamento in pagamentos:
            origem = pagamento.registrado_por or "?"
            resumo["por_origem"][origem] = resumo["por_origem"].get(origem, servico.ZERO) + servico.dinheiro(pagamento.valor)

        alerta = "  <-- em mais de uma turma" if len(turmas) > 1 else ""
        relatar(f"{aluno.nome} (#{aluno.cod_alu}){alerta}")
        for cod_tur in turmas:
            do_turma = [c for c in cobrancas if c.cod_tur == cod_tur]
            primeira = [c for c in do_turma if c.parcela == 1 and c.tipo != "AVULSA"]
            valor_primeira = sum((servico.dinheiro(c.valor) for c in primeira), servico.ZERO)
            pago_primeira = sum((pagos.get(c.id, servico.ZERO) for c in primeira), servico.ZERO)
            relatar(
                f"    turma {cod_tur} {nomes_turma.get(cod_tur, '?'):<12} "
                f"{len(do_turma):>3} cobrança(s) · parcela 1 soma {valor_primeira} "
                f"(pago {pago_primeira}) · criadas por "
                f"{', '.join(sorted({c.criado_por or '?' for c in do_turma}))}"
            )
        if pagamentos:
            por_quem = {}
            for pagamento in pagamentos:
                quem = pagamento.registrado_por or "?"
                por_quem[quem] = por_quem.get(quem, servico.ZERO) + servico.dinheiro(pagamento.valor)
            detalhe = ", ".join(f"{quem} R$ {valor}" for quem, valor in sorted(por_quem.items()))
            relatar(f"    baixas: R$ {total_pago} em {len(pagamentos)} lançamento(s) — {detalhe}")

    relatar("")
    relatar("Resumo")
    relatar(f"  alunos com cobrança já criada ....... {resumo['com_cobranca']}")
    relatar(f"  alunos com cobrança em duas turmas .. {resumo['em_mais_de_uma_turma']}")
    relatar(f"  alunos com baixa lançada ............ {resumo['com_pagamento']}")
    for origem, valor in sorted(resumo["por_origem"].items()):
        relatar(f"    baixas registradas por {origem}: R$ {valor}")
    if resumo["em_mais_de_uma_turma"]:
        relatar("")
        relatar(
            "  Cobrança em duas turmas duplica o que o aluno deve: a chave é "
            "(aluno, turma, tipo, parcela), então o mesmo mês existe duas vezes. "
            "Use --limpar-importacao para desfazer o que a planilha criou antes."
        )
    return resumo


def limpar_importacao(db, correspondencias: list[dict], relatar=print) -> dict:
    """Desfaz o que uma importação anterior da planilha criou.

    Remove só o que tem a marca ``PLANILHA``: as baixas que ela lançou e as
    cobranças que ela criou e que não receberam nenhum outro pagamento. O que
    a secretaria lançou na tela fica de pé — o script não apaga trabalho de
    gente.
    """
    codigos = [item["aluno"].cod_alu for item in correspondencias if item.get("aluno")]
    if not codigos:
        return {"pagamentos": 0, "cobrancas": 0, "preservadas": 0}

    cobrancas = list(
        db.scalars(select(Cobranca).where(Cobranca.cod_alu.in_(codigos)))
    )
    ids = [c.id for c in cobrancas]
    pagamentos = list(
        db.scalars(select(Pagamento).where(Pagamento.cobranca_id.in_(ids)))
    ) if ids else []

    removidos = 0
    for pagamento in pagamentos:
        if pagamento.registrado_por == ORIGEM:
            db.delete(pagamento)
            removidos += 1
    db.flush()

    apagadas = preservadas = 0
    for cobranca in cobrancas:
        if cobranca.criado_por != ORIGEM:
            continue
        if servico.total_pago(db, cobranca.id) > servico.ZERO:
            preservadas += 1
            continue
        db.delete(cobranca)
        apagadas += 1
    db.flush()
    relatar(
        f"Limpeza: {removidos} baixa(s) e {apagadas} cobrança(s) da importação removidas; "
        f"{preservadas} cobrança(s) preservada(s) por ter pagamento de outra origem."
    )
    return {"pagamentos": removidos, "cobrancas": apagadas, "preservadas": preservadas}


def resumir_para_sobrescrever(db, correspondencias: list[dict]) -> dict:
    """Conta o que a sobrescrita apagaria, sem apagar nada."""
    codigos = [item["aluno"].cod_alu for item in correspondencias if item.get("aluno")]
    if not codigos:
        return {"cobrancas": 0, "pagamentos": 0, "valor": servico.ZERO, "avulsas": 0}
    cobrancas = list(
        db.scalars(
            select(Cobranca).where(
                Cobranca.cod_alu.in_(codigos),
                Cobranca.tipo.in_(("MATRICULA", "MENSALIDADE")),
            )
        )
    )
    ids = [c.id for c in cobrancas]
    pagamentos = (
        list(db.scalars(select(Pagamento).where(Pagamento.cobranca_id.in_(ids)))) if ids else []
    )
    avulsas = db.scalar(
        select(func.count())
        .select_from(Cobranca)
        .where(Cobranca.cod_alu.in_(codigos), Cobranca.tipo == "AVULSA")
    ) or 0
    return {
        "cobrancas": len(cobrancas),
        "pagamentos": len(pagamentos),
        "valor": sum((servico.dinheiro(p.valor) for p in pagamentos), servico.ZERO),
        "avulsas": int(avulsas),
    }


def sobrescrever(db, correspondencias: list[dict], relatar=print) -> dict:
    """Apaga matrícula e mensalidade destes alunos para reconstruir da planilha.

    Diferente de ``limpar_importacao``, aqui não importa quem criou: some tudo
    que veio do plano, inclusive baixa lançada na tela. É o que a planilha
    pede quando ela é a fonte da verdade e o banco tem sobra de tentativas
    anteriores. Cobrança avulsa fica de pé: ela nunca veio do plano.

    Recebimento bancário conciliado com uma cobrança que vai embora volta para
    a fila de conciliação — o dinheiro entrou no banco de todo jeito.
    """
    resumo = resumir_para_sobrescrever(db, correspondencias)
    codigos = [item["aluno"].cod_alu for item in correspondencias if item.get("aluno")]
    if not codigos:
        return resumo

    cobrancas = list(
        db.scalars(
            select(Cobranca).where(
                Cobranca.cod_alu.in_(codigos),
                Cobranca.tipo.in_(("MATRICULA", "MENSALIDADE")),
            )
        )
    )
    ids = [c.id for c in cobrancas]
    if ids:
        for pagamento in db.scalars(select(Pagamento).where(Pagamento.cobranca_id.in_(ids))):
            if pagamento.transacao_id:
                transacao = db.get(TransacaoBancaria, pagamento.transacao_id)
                if transacao:
                    transacao.status = "PENDENTE"
                    transacao.cobranca_id = None
                    transacao.motivo = "Cobrança removida na reimportação da planilha"
                    transacao.conciliada_em = None
                    transacao.conciliada_por = None
            db.delete(pagamento)
        db.flush()
    for cobranca in cobrancas:
        db.delete(cobranca)
    db.flush()
    relatar(
        f"Sobrescrita: {resumo['cobrancas']} cobrança(s) e {resumo['pagamentos']} baixa(s) "
        f"removidas (R$ {resumo['valor']}); {resumo['avulsas']} avulsa(s) preservada(s)."
    )
    return resumo


# ---- orquestração ----------------------------------------------------------

def importar(
    db,
    importados: list[dict],
    totais: dict,
    *,
    parcelas: int,
    matricula: Decimal,
    mensalidade: Decimal,
    desconto_conjuge: Decimal,
    primeira_mensalidade: date,
    dia_vencimento: int = 10,
    parcelas_transferencia: int | None = None,
    relatar=print,
) -> dict:
    """Grava no banco e devolve o que foi feito. Não faz commit."""
    # 1. Plano em cada turma que tem gente na planilha.
    turmas = sorted({item["cod_tur"] for item in importados})
    planos = {
        cod_tur: garantir_plano(
            db,
            cod_tur,
            matricula=matricula,
            mensalidade=mensalidade,
            parcelas=parcelas,
            primeira_mensalidade=primeira_mensalidade,
            dia_vencimento=dia_vencimento,
        )
        for cod_tur in turmas
    }
    relatar(f"Plano aplicado a {len(turmas)} turma(s): {', '.join(str(t) for t in turmas)}")

    # 2. Aluno, matrícula na turma e condição.
    criados = 0
    for item in importados:
        aluno = item.get("aluno")
        if aluno is None:
            aluno = Aluno(
                nome=item["nome"],
                status="A",
                dat_cad=date.today(),
                origem_cadastro=ORIGEM,
            )
            db.add(aluno)
            db.flush()
            item["aluno"] = aluno
            criados += 1
        if turma_do_aluno(db, aluno) != item["cod_tur"]:
            sincronizar_matricula(db, aluno, item["cod_tur"])
        definir_condicao(
            db,
            item,
            aluno,
            item["cod_tur"],
            desconto_conjuge=desconto_conjuge,
            parcelas_transferencia=parcelas_transferencia,
        )
    db.flush()
    relatar(f"Alunos: {criados} criado(s), {len(importados) - criados} já cadastrado(s).")

    # 3. Cobranças de quem está na planilha — e só dele.
    #    A geração da turma inteira cobraria também quem a planilha não cita.
    cobrancas_criadas = transferencias = 0
    nomes_turma = {cod: nome for cod, nome in db.execute(select(Turma.cod_tur, Turma.nome))}
    for item in importados:
        plano = planos[item["cod_tur"]]
        condicao = db.scalar(
            select(CondicaoFinanceiraAluno).where(
                CondicaoFinanceiraAluno.cod_tur == item["cod_tur"],
                CondicaoFinanceiraAluno.cod_alu == item["aluno"].cod_alu,
            )
        )
        efetivo = servico.plano_efetivo(plano, condicao)
        if efetivo.transferencia:
            transferencias += 1
        resultado = servico.aplicar_plano_ao_aluno(
            db,
            plano,
            item["aluno"].cod_alu,
            efetivo=efetivo,
            nome_turma=nomes_turma.get(item["cod_tur"]) or f"Turma {item['cod_tur']}",
            criado_por=ORIGEM,
        )
        cobrancas_criadas += resultado["criadas"]
    relatar(
        f"Cobranças: {cobrancas_criadas} criada(s), "
        f"{transferencias} aluno(s) com condição de transferência."
    )

    # 4. Baixas do que já foi pago.
    por_linha = {item["linha"]: item for item in importados}
    baixado = servico.ZERO
    for item in importados:
        if item["conjuge"] or item["pago"] <= servico.ZERO:
            continue
        alunos_do_pagamento = [item["aluno"]]
        if item["titular_casal"] and (item["linha"] + 1) in por_linha:
            alunos_do_pagamento.append(por_linha[item["linha"] + 1]["aluno"])
        baixado += quitar(
            db, item, alunos_do_pagamento, data_pagamento=primeira_mensalidade, relatar=relatar
        )
    relatar(f"Baixas: R$ {baixado} lançado(s).")
    db.flush()

    return {
        "turmas": turmas,
        "alunos_criados": criados,
        "alunos_reaproveitados": len(importados) - criados,
        "cobrancas_criadas": cobrancas_criadas,
        "transferencias": transferencias,
        "baixado": baixado,
        "confere": conferir(db, importados, totais, relatar=relatar),
    }


def relatar_correspondencia(correspondencias: list[dict], db) -> None:
    """Mostra, nome por nome, o que o script encontrou no cadastro."""
    rotulos = {
        "EXATO": "exato",
        "PARCIAL": "INCOMPLETO",
        "APROXIMADO": "PARECIDO",
        "AMBIGUO": "AMBÍGUO",
        "NOVO": "não cadastrado",
    }
    nomes_turma = {
        cod: nome for cod, nome in db.execute(select(Turma.cod_tur, Turma.nome))
    }
    print()
    print(f"{'PLANILHA':<38} {'SITUAÇÃO':<15} CADASTRO / TURMA")
    print("-" * 100)
    for item in correspondencias:
        detalhe = ""
        if item["situacao"] in ("EXATO", "PARCIAL", "APROXIMADO"):
            aluno = item["aluno"]
            cod_tur = turma_do_aluno(db, aluno)
            turma = nomes_turma.get(cod_tur, "sem turma")
            detalhe = f"{aluno.nome} — {turma}"
            if item["situacao"] == "APROXIMADO":
                detalhe += f"  ({item['semelhanca']:.0%} de semelhança)"
        elif item["situacao"] == "AMBIGUO":
            detalhe = " | ".join(f"{a.nome} (#{a.cod_alu})" for a in item["candidatos"][:3])
        print(f"{item['nome'][:37]:<38} {rotulos[item['situacao']]:<15} {detalhe}")


def main() -> None:
    analisador = argparse.ArgumentParser(description=__doc__)
    analisador.add_argument("--arquivo", required=True, help="Caminho do .xlsx")
    analisador.add_argument(
        "--parcelas", type=int, default=24, help="Mensalidades do curso (2 anos = 24)"
    )
    analisador.add_argument(
        "--parcelas-transferencia",
        type=int,
        help="Mensalidades de quem veio por transferência (padrão: igual à turma)",
    )
    analisador.add_argument(
        "--turma-novos",
        type=int,
        help="Turma para quem não está no cadastro ou está sem turma",
    )
    analisador.add_argument("--matricula", type=Decimal, default=Decimal("100"))
    analisador.add_argument("--mensalidade", type=Decimal, default=Decimal("200"))
    analisador.add_argument("--desconto-conjuge", type=Decimal, default=Decimal("50"))
    analisador.add_argument("--primeira-mensalidade", default="2026-08-10")
    analisador.add_argument("--dia-vencimento", type=int, default=10)
    analisador.add_argument(
        "--aceitar-aproximados",
        action="store_true",
        help="Aceita nome incompleto ou parecido quando há um único candidato",
    )
    analisador.add_argument(
        "--criar-novos", action="store_true", help="Cadastra quem não foi encontrado"
    )
    analisador.add_argument(
        "--diagnostico",
        action="store_true",
        help="Mostra o que já existe no banco para estes alunos e sai",
    )
    analisador.add_argument(
        "--limpar-importacao",
        action="store_true",
        help="Desfaz o que uma importação anterior da planilha criou (exige --aplicar)",
    )
    analisador.add_argument(
        "--sobrescrever",
        action="store_true",
        help=(
            "A planilha vira a verdade: apaga matrícula e mensalidade destes alunos, "
            "inclusive baixa lançada na tela, e reconstrói (exige --aplicar)"
        ),
    )
    analisador.add_argument(
        "--aplicar",
        action="store_true",
        help="Grava de verdade; sem esta opção o script só mostra o que faria",
    )
    args = analisador.parse_args()

    Base.metadata.create_all(engine)
    atualizar_schema(engine)
    db = SessionLocal()
    try:
        pessoas, totais = ler_planilha(args.arquivo)
        correspondencias = casar_nomes(db, pessoas)
        vencimento = date.fromisoformat(args.primeira_mensalidade)

        print(f"Planilha: {args.arquivo}")
        print(
            f"  {len(pessoas)} pessoa(s): "
            f"{sum(1 for p in pessoas if p['bloco'] == 'REGULAR')} regular(es), "
            f"{sum(1 for p in pessoas if p['bloco'] == 'TRANSFERENCIA')} transferido(s), "
            f"{sum(1 for p in pessoas if p['conjuge'])} cônjuge(s)"
        )
        contagem = {}
        for item in correspondencias:
            contagem[item["situacao"]] = contagem.get(item["situacao"], 0) + 1
        print("  correspondência: " + ", ".join(f"{v} {k.lower()}" for k, v in sorted(contagem.items())))
        relatar_correspondencia(correspondencias, db)

        if args.diagnostico:
            diagnosticar(db, correspondencias)
            return

        if args.limpar_importacao:
            if not args.aplicar:
                diagnosticar(db, correspondencias)
                print("\nSimulação: repita com --aplicar para remover.")
                return
            limpar_importacao(db, correspondencias)
            db.commit()
            print("\nLimpeza concluída. Rode a importação de novo.")
            return

        entram, de_fora = selecionar(
            db,
            correspondencias,
            aceitar_aproximados=args.aceitar_aproximados,
            criar_novos=args.criar_novos,
            turma_novos=args.turma_novos,
        )
        if de_fora:
            print()
            print("Ficam de fora:")
            for item, motivo in de_fora:
                print(f"  - {item['nome']}: {motivo}")
        if not entram:
            print("\nNinguém a importar. Turmas cadastradas:")
            for turma in db.scalars(select(Turma).order_by(Turma.nome)):
                print(f"  {turma.cod_tur:>3}  {turma.nome}")
            return

        if args.sobrescrever:
            aviso = resumir_para_sobrescrever(db, [i for i in entram])
            print()
            print(
                f"SOBRESCREVER apagaria {aviso['cobrancas']} cobrança(s) e "
                f"{aviso['pagamentos']} baixa(s), somando R$ {aviso['valor']}, "
                f"para reconstruir tudo a partir da planilha."
            )
            print(f"  {aviso['avulsas']} cobrança(s) avulsa(s) ficam de pé.")

        turmas = sorted({item["cod_tur"] for item in entram})
        print()
        print(
            f"Entram {len(entram)} pessoa(s) em {len(turmas)} turma(s). "
            f"Plano: matrícula R$ {args.matricula}, mensalidade R$ {args.mensalidade}, "
            f"{args.parcelas} parcelas, 1ª em {vencimento:%d/%m/%Y}."
        )
        if not args.aplicar:
            print("\nSimulação: nada foi gravado. Repita com --aplicar para valer.")
            return

        if args.sobrescrever:
            sobrescrever(db, entram)

        resultado = importar(
            db,
            entram,
            totais,
            parcelas=args.parcelas,
            matricula=args.matricula,
            mensalidade=args.mensalidade,
            desconto_conjuge=args.desconto_conjuge,
            primeira_mensalidade=vencimento,
            dia_vencimento=args.dia_vencimento,
            parcelas_transferencia=args.parcelas_transferencia,
        )
        if not resultado["confere"]:
            db.rollback()
            sys.exit(
                "\nNada foi gravado: o resultado não bate com a planilha.\n"
                "Se o banco já tinha cobrança ou baixa destes alunos, veja com "
                "--diagnostico; para a planilha valer sobre o que existe, use "
                "--sobrescrever --aplicar."
            )
        db.commit()
        print("\nImportação concluída.")
        if resultado["transferencias"]:
            print(
                "Revise quantos módulos cada transferido ainda vai cursar em "
                "Financeiro › turma › Condição."
            )
    finally:
        db.close()


if __name__ == "__main__":
    main()
