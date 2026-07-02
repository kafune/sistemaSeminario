import csv
import io
import zipfile

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import Response
from sqlalchemy import select
from sqlalchemy.orm import Session

from ..database import get_db
from ..models import Aluno, AluTurma
from ..pdf.boletim import gerar_boletim
from ..pdf.diario import gerar_diario
from ..pdf.historico import gerar_historico
from ..pdf.listas import gerar_ficha_aluno, gerar_lista_turma

router = APIRouter(prefix="/relatorios", tags=["relatorios"])


def _pdf(conteudo: bytes, nome: str) -> Response:
    return Response(
        content=conteudo,
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{nome}"'},
    )


@router.get("/boletim/{cod_alu}")
def boletim(cod_alu: int, db: Session = Depends(get_db)):
    try:
        return _pdf(gerar_boletim(db, cod_alu), f"Boletim_{cod_alu}.pdf")
    except ValueError as e:
        raise HTTPException(404, str(e))


@router.get("/boletim-turma/{cod_tur}")
def boletins_da_turma(cod_tur: int, db: Session = Depends(get_db)):
    """ZIP com o boletim de cada aluno da turma (substitui o /criartodos antigo)."""
    alunos = list(
        db.execute(
            select(Aluno.cod_alu, Aluno.nome)
            .join(AluTurma, AluTurma.cod_alu == Aluno.cod_alu)
            .where(AluTurma.cod_tur == cod_tur)
            .order_by(Aluno.nome)
        )
    )
    if not alunos:
        raise HTTPException(404, "Turma sem alunos matriculados")
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for cod_alu, nome in alunos:
            nome_arquivo = "".join(c for c in (nome or str(cod_alu)) if c.isalnum() or c in " _-")
            zf.writestr(f"Boletim_{nome_arquivo}.pdf", gerar_boletim(db, cod_alu))
    return Response(
        content=buffer.getvalue(),
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="boletins_turma_{cod_tur}.zip"'
        },
    )


# ---- geração em lote a partir de arquivo CSV/XLSX/XLS ----------------------

CABECALHOS_IGNORADOS = {"matricula", "matrícula", "cod_alu", "codigo", "código", "nome", "aluno", "alunos"}


def _primeira_coluna(nome_arquivo: str, conteudo: bytes) -> list[str]:
    """Extrai os valores da primeira coluna do arquivo enviado."""
    nome = (nome_arquivo or "").lower()
    valores: list[str] = []
    if nome.endswith(".csv") or nome.endswith(".txt"):
        try:
            texto = conteudo.decode("utf-8-sig")
        except UnicodeDecodeError:
            texto = conteudo.decode("latin-1")
        # aceita ; ou , como separador (Excel brasileiro usa ;)
        amostra = texto[:2048]
        sep = ";" if amostra.count(";") >= amostra.count(",") else ","
        for linha in csv.reader(io.StringIO(texto), delimiter=sep):
            if linha and str(linha[0]).strip():
                valores.append(str(linha[0]).strip())
    elif nome.endswith(".xlsx"):
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
        ws = wb.worksheets[0]
        for row in ws.iter_rows(values_only=True):
            if row and row[0] is not None and str(row[0]).strip():
                valores.append(str(row[0]).strip())
        wb.close()
    elif nome.endswith(".xls"):
        import xlrd

        wb = xlrd.open_workbook(file_contents=conteudo)
        ws = wb.sheet_by_index(0)
        for i in range(ws.nrows):
            v = ws.cell_value(i, 0)
            if v is not None and str(v).strip():
                valores.append(str(v).strip())
    else:
        raise HTTPException(400, "Formato não suportado. Envie um arquivo .csv, .xlsx ou .xls")
    return [v for v in valores if v.lower() not in CABECALHOS_IGNORADOS]


def _achar_aluno(db: Session, valor: str) -> Aluno | None:
    """Localiza o aluno por matrícula (número) ou por nome."""
    v = valor.strip()
    # planilhas costumam guardar números como '123.0'
    if v.replace(".0", "").isdigit():
        return db.get(Aluno, int(float(v)))
    aluno = db.scalar(select(Aluno).where(Aluno.nome == v).limit(1))
    if aluno:
        return aluno
    candidatos = list(db.scalars(select(Aluno).where(Aluno.nome.like(f"%{v}%")).limit(2)))
    # só aceita busca parcial quando o resultado é único (evita pegar homônimos)
    return candidatos[0] if len(candidatos) == 1 else None


@router.post("/lote")
def relatorios_em_lote(
    tipo: str = Query("boletim", pattern="^(boletim|historico)$"),
    arquivo: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """Recebe um CSV/XLSX/XLS com matrículas ou nomes na primeira coluna e
    devolve um ZIP com o boletim ou histórico de cada aluno encontrado."""
    valores = _primeira_coluna(arquivo.filename, arquivo.file.read())
    if not valores:
        raise HTTPException(400, "O arquivo não tem valores na primeira coluna")

    gerar = gerar_boletim if tipo == "boletim" else gerar_historico
    prefixo = "Boletim" if tipo == "boletim" else "Historico"

    nao_encontrados: list[str] = []
    gerados = 0
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        vistos: set[int] = set()
        for valor in valores:
            aluno = _achar_aluno(db, valor)
            if not aluno:
                nao_encontrados.append(valor)
                continue
            if aluno.cod_alu in vistos:
                continue
            vistos.add(aluno.cod_alu)
            nome_arquivo = "".join(
                c for c in (aluno.nome or str(aluno.cod_alu)) if c.isalnum() or c in " _-"
            )
            zf.writestr(f"{prefixo}_{nome_arquivo}.pdf", gerar(db, aluno.cod_alu))
            gerados += 1
        if nao_encontrados:
            zf.writestr(
                "_NAO_ENCONTRADOS.txt",
                "Estes valores não foram localizados no cadastro de alunos:\r\n\r\n"
                + "\r\n".join(nao_encontrados),
            )
    if gerados == 0:
        raise HTTPException(404, "Nenhum aluno do arquivo foi encontrado no cadastro")
    return Response(
        content=buffer.getvalue(),
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{prefixo.lower()}s_lote.zip"'},
    )


@router.get("/historico/{cod_alu}")
def historico(cod_alu: int, db: Session = Depends(get_db)):
    try:
        return _pdf(gerar_historico(db, cod_alu), f"Historico_{cod_alu}.pdf")
    except ValueError as e:
        raise HTTPException(404, str(e))


@router.get("/diario/{cod_tur}")
def diario(cod_tur: int, cod_mat: int | None = None, db: Session = Depends(get_db)):
    try:
        return _pdf(gerar_diario(db, cod_tur, cod_mat), f"Diario_{cod_tur}.pdf")
    except ValueError as e:
        raise HTTPException(404, str(e))


@router.get("/lista-turma/{cod_tur}")
def lista_turma(cod_tur: int, db: Session = Depends(get_db)):
    try:
        return _pdf(gerar_lista_turma(db, cod_tur), f"Alunos_Turma_{cod_tur}.pdf")
    except ValueError as e:
        raise HTTPException(404, str(e))


@router.get("/ficha-aluno/{cod_alu}")
def ficha_aluno(cod_alu: int, db: Session = Depends(get_db)):
    try:
        return _pdf(gerar_ficha_aluno(db, cod_alu), f"Ficha_{cod_alu}.pdf")
    except ValueError as e:
        raise HTTPException(404, str(e))
